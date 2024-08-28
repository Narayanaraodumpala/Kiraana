from django.shortcuts import render
from django.views.generic import View
from django.http import HttpResponse
from product.models import Category,Brands
import random
import string
from django.contrib import messages
from django.views.generic.list import ListView
from django.shortcuts import redirect

# Create your views here.

def random_string(length=8, uppercase=True,
                  lowecase=True, numbers=True):
    charecter_set = '-'
    if uppercase:
        charecter_set += string.ascii_uppercase
    if lowecase:
        charecter_set += string.ascii_lowercase
    if numbers:
        charecter_set += string.digits
    return ''.join(random.choice(charecter_set) for i in range(length))


my_random = random_string(8)


class Savecategory(View):
    def get(self, request, *args, **kwargs):
        return render(request,'superadmin/addcategory.html')

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
           
            cimage=request.FILES['cimage']
            cname=request.POST['cname']
            ctype=request.POST['ctype']
            cdescription=request.POST['cdescription']
            category_code='#C1'+my_random
            category=Category.objects.filter(category_name=cname)
            if not  category:
                Category.objects.create(category_name=cname,category_code=category_code,
                                            description=cdescription,category_type=ctype,category_image=cimage)
                messages.success(request, 'Category Uploaded Successfully!')
            
                return render(request,'superadmin/addcategory.html')
            else:
                messages.error(request, 'Category  Already Uploaded!')
            
                return render(request,'superadmin/addcategory.html')
        else:
            return render(request,'superadmin/addcategory.html')
        
class Display_categories(ListView):
    model = Category
    template_name = 'superadmin/categorylist.html'
    
import os   
class  Editcategory(View):
    def get(self,request,pk,*args, **kwargs):
    
        cat=Category.objects.filter(id=pk)
        
        return render(request,'superadmin/editcategory.html',{'category':cat})
    
    def post(self,request,pk,*args, **kwargs):
        category = Category.objects.get(id=pk)
        if request.method == 'POST':
        
            
            if len(request.FILES) != 0:
              if len(category.category_image) > 0:
                os.remove(category.category_image.path)
                category.category_image = request.FILES['cimage']
                category.category_name = request.POST.get('cname')
                category.description = request.POST.get('cdescription')
                category.category_type = request.POST.get('ctype')
                category.save()
                
                return redirect('display_categories')
            else:
                pass
                
            
def deletecategory(request,pk):
    category = Category.objects.filter(id=pk)
    return render(request,'superadmin/deletecategory.html',{'category':category})


def deletecategorydone(request,pk):
    category = Category.objects.filter(id=pk)
    if category:
        category.delete()
        return redirect('display_categories')
    else:
        return redirect('display_categories')
    
    
def deletebrand(request,pk):
    brand = Brands.objects.filter(id=pk)
    return render(request,'superadmin/deletebrand.html',{'brand':brand})


def deletebranddone(request,pk):
    brand = Brands.objects.filter(id=pk)
    if brand:
        brand.delete()
        return redirect('brandlist')
    else:
        return redirect('brandlist')
    
class Addbrand(View):
    def get(self,request,*args, **kwargs):
        return render(request,'superadmin/addbrands.html')
    
    def post(self,request,*args, **kwargs):
        if request.method == 'POST':
           
            bimage=request.FILES['bimage']
            bname=request.POST['bname']
            
            bdescription=request.POST['bdescription']
            brand_code='#B1'+my_random
            brand=Brands.objects.filter(brand_name=bname)
            if not  brand:
                Brand.objects.create(brand_name=bname,brand_code=brand_code,
                                            description=bdescription,brand_image=bimage)
                messages.success(request, 'Brand Uploaded Successfully!')
            
                return render(request,'superadmin/addbrands.html')
            else:
                messages.error(request, 'Brand  Already Uploaded!')
            
                return render(request,'superadmin/addbrands.html')
        else:
            return render(request,'superadmin/addbrands.html')
        
        
        
class BrandList(ListView):
    model = Brands
    template_name = 'superadmin/brandlist.html'
        
        
class Editbrand(View):
    def get(self,request,pk,*args, **kwargs):
    
        brand=Brands.objects.filter(id=pk)
        
        return render(request,'superadmin/editbrand.html',{'brand':brand})
    
    def post(self,request,pk,*args, **kwargs):
        brand = Brands.objects.get(id=pk)
        print('brand=',brand)
        if len(request.FILES) != 0:
              print('----')
              if len(brand.brand_image) > 0:
                
                
                os.remove(brand.brand_image.path)
                brand.brand_image = request.FILES['bimage']
                brand.brand_name = request.POST.get('bname')
                brand.description = request.POST.get('bdescription')
               
                brand.save()
                
                return redirect('brandlist')
        else:
                pass
            
            
            
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A2
from reportlab.lib.units import inch
from reportlab.lib.colors import black, white, blue, red
from datetime import date
from reportlab.lib import colors

from django.template.loader import get_template

from xhtml2pdf import pisa            
       

def categoryexport_pdf(request):
    category = Category.objects.all().order_by('id')

    template_path = 'pdf_convert/CategoryPdfReport.html'

    context = {'categories': category}

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'filename="CategoryInfo_Report.pdf"'

    template = get_template(template_path)

    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse

def categoryexport_excel(request):

    categories = Category.objects.all()

    # Create a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set header row
    header_row = ['Category Name', 'Categor Code','Category Type','Description', 'Category Image']
    for col, value in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Iterate over users and add data to Excel file
    row_num = 2
    for category in categories:
        row_data = [
            category.category_name,
            category.category_code,
            category.category_type,
            category.description,
            category.category_image.url,
           
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
        row_num += 1

    # Save the Excel file to a BytesIO object
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return the Excel file as a response
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Category.xlsx"'
    return response


def brandexport_pdf(request):
    brand = Brands.objects.all().order_by('id')

    template_path = 'pdf_convert/BrandPdfReport.html'

    context = {'brand': brand}

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'filename="BrandInfo_Report.pdf"'

    template = get_template(template_path)

    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response



def brandexport_excel(request):
    brand = Brands.objects.all()

    # Create a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set header row
    header_row = ['Brand Name', 'Brand Code','Description', 'Brand Image']
    for col, value in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Iterate over users and add data to Excel file
    row_num = 2
    for category in brand:
        row_data = [
            category.brand_name,
            category.brand_code,
            
            category.description,
            category.brand_image.url,
           
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
        row_num += 1

    # Save the Excel file to a BytesIO object
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return the Excel file as a response
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Brand.xlsx"'
    return response
            
            
           
                

    
 