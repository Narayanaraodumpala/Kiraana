from django.shortcuts import render

# Create your views here.
from django.shortcuts import render, redirect

from . forms import CreateUserForm, LoginForm
from product.models import Category,Brands
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.views.generic import View
from django.db import IntegrityError
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from django.urls import reverse
from django.conf import settings

import socket
from app1.models import UserData
from django.contrib.sites.shortcuts import get_current_site
from datetime import timedelta
import smtplib
import random
import string
import jwt
import uuid
from .helpers import send_forgot_password_mail

import json


# - Authentication models and functions

from django.contrib.auth.models import auth
from django.contrib.auth import authenticate, login, logout


if __name__ == '__main__':
    pass

sender_address = 'spnnandu@gmail.com'
sender_pass = 'yfim xazb euqb rvux'

socket.getaddrinfo('localhost', 8080)

# Create your views here.

def random_string(length=8, uppercase=True,
                  lowecase=True, numbers=True):
    charecter_set = '-'
    if uppercase:
        charecter_set += string.ascii_uppercase
    if lowecase:
        charecter_set += string.ascii_lowercase
    if numbers:
        charecter_set += string.digits
    return ''.join(random.choice(charecter_set) for i in range(length))


my_random = random_string(8)

# Create your views here.

def index(request):
    return render(request,'user/index.html')



def superadminview(request):
    users_count=UserData.objects.all().count()
    users=UserData.objects.all()
    category=Category.objects.all()
    brand=Brands.objects.all()
    totalcategory=Category.objects.all().count()
    totalbrand=Brands.objects.all().count()
    
    
    return render(request,'superadmin/index.html',{'user_count':users_count,
                                                   'users':users,'category':category,'brand':brand,
                                                   'totalcategory':totalcategory,'totalbrand':totalbrand})

class VerifyEmail(View):
    
    def get(self, request):
        token = request.GET.get('token')
        
        try:
            payload = jwt.decode(
                token, settings.SECRET_KEY, algorithms=["HS256"])
            
            

            user = User.objects.get(id=payload['user_id'])
            if not user.is_active:
                user.is_active = True
                user.save()

            return render(request, 'user/signup.html')
        except jwt.ExpiredSignatureError as identifier:
            return render(request, 'user/signup.html')


class Rigester(View):
    
    def get(self,request):
        
        return render(request,'user/login.html')
    
    def post(self,request):

        if request.method == "POST":
            try:
                username=request.POST['username']
                email=request.POST['email']
                password=request.POST['password']
            # User.objects.create(username=username,email=email,password=password)
                is_active = '0'
                unique_user_id = '#1'+my_random
                print('unique_user_id=', unique_user_id)
                usr = User.objects.filter(email=email)
                if not usr:
                            user = User.objects.create_user(
                                username=username, password=password, email=email, is_active=is_active)
                            UserData.objects.create(
                                user=user,  unique_user_id=unique_user_id,)
                            user = User.objects.get(email=email)
                            token = RefreshToken.for_user(user).access_token
                            token.set_exp(lifetime=timedelta(days=36500))
                            current_site = get_current_site(request).domain
                            relativeLink = reverse('email-verify')
                            absurl = 'http://' + current_site + relativeLink +'?token=' + str(token)
                            reciver_mail = user.email
                            message = MIMEMultipart()
                            message['From'] = sender_address
                            message['To'] = reciver_mail
                            message['Subject'] = 'Registration confirmation! '
                            mail_content = 'hello' + ' ' + user.username + ' please click this below  link to verify your account ' \
                                '\n ' + absurl
                            message.attach(MIMEText(mail_content, 'plain'))
                            s = smtplib.SMTP('smtp.gmail.com', 587)
                            s.starttls()
                            s.login(sender_address, sender_pass)
                            text = message.as_string()
                            s.sendmail(sender_address, reciver_mail, text)
                            return render(request, 'user/signup.html',{'error':'An Email is sent for Verifying to '+ user.email+ ', Please verify and login'})
                else:
                            
                 return render(request, 'user/login.html',{'error':'Account With This User is Already registered'})
            except IntegrityError as e:
                return  render(request, 'user/login.html')
        

class Signin(View):
    def get(self,request):
        return render(request,'user/signup.html')
    
    
    def post(self,request):
        # if request.user.is_authenticated:
        #          return redirect('index')
        # else:
        #      error = False
        if request.method =='POST':
           
            dic = request.POST
            usr = dic['username']
            pwd = dic['password']
            data = User.objects.filter(username=usr).first()
           
            if not data:
                
                message = 'incorrect username'
                print(message)
                return render(request, 'user/signup.html', {'user_mesg': message})
            else:

                if not data.is_active:
                    message = 'account not verified, Please Check your mail to Activate the account'
                    print(message)
                    return render(request, 'user/signup.html', {'message': message})

                else:
                    user = authenticate(username=usr, password=pwd)
                    if user:
                        login(request, user)
                        data = User.objects.filter(username=request.user).first()
                       
                        return redirect('home')
                    else:
                        error = 'invalid password'
                        return render(request, 'user/signup.html', {'error': error})


def user_logout(request):
    
   
    logout(request)
    
        
    return redirect('home') 


     
def request_reset(request):
    try:
        if request.method == 'POST':
            username = request.POST.get('username')
           

            if not User.objects.filter(username=username).first():

                message = 'sorry , we are not found record with this username'
                return render(request, 'user/request_reset.html', {'message': message})
            else:

                user_obj = User.objects.get(username=username)

                token = str(uuid.uuid4())
                profile_obj = UserData.objects.get(user=user_obj)
                profile_obj.forgot_password_token = token

                profile_obj.save()

             
                send_forgot_password_mail(user_obj.email, token, request)

                print('email sent')
                message = 'email sent'
                return render(request, 'user/request_reset.html', {'message': message})
    except Exception as e:
        print(e)
        
    return render(request,'user/request_reset.html')


def profile(request):
    return render(request,'user/Dashboard/index2.html')


import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse


def export_list(request):
    # Get all users
    users = UserData.objects.all()

    # Create a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set header row
    header_row = ['Username', 'Email','unique_user_id','forgot_password_token', 'First Name', 'Last Name']
    for col, value in enumerate(header_row, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = value
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Iterate over users and add data to Excel file
    row_num = 2
    for user in users:
        row_data = [
            user. user.username,
            user. user.email,
            user.unique_user_id,
            user.forgot_password_token,
            user. user.first_name,
            user. user.last_name,
        ]
        for col, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col)
            cell.value = value
        row_num += 1

    # Save the Excel file to a BytesIO object
    from io import BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Return the Excel file as a response
    response = HttpResponse(
        excel_file.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
    return response

from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A2
from reportlab.lib.units import inch
from reportlab.lib.colors import black, white, blue, red
from datetime import date
from reportlab.lib import colors

from django.template.loader import get_template

from xhtml2pdf import pisa

def export_pdf(request):
    
    
    
    users = UserData.objects.all().order_by('id')

    template_path = 'pdf_convert/pdfReport.html'

    context = {'users': users}

    response = HttpResponse(content_type='application/pdf')

    response['Content-Disposition'] = 'filename="UsersInfo_report.pdf"'

    template = get_template(template_path)

    html = template.render(context)

    # create a pdf
    pisa_status = pisa.CreatePDF(
       html, dest=response)
    # if error then show some funy view
    if pisa_status.err:
       return HttpResponse('We had some errors <pre>' + html + '</pre>')
    return response


def import_excel(request):
    users_count=UserData.objects.all().count()
    return render(request,'superadmin/import_excel.html',{'user_count':users_count})

import pandas as pd
from django.core.files.storage import FileSystemStorage
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def import_excel_file(request):
      if request.method == 'POST' and request.FILES['myfile']:
            myfile = request.FILES['myfile']
            #fs = FileSystemStorage()
            #filename = fs.save(myfile.name, myfile)
            # uploaded_file_url = fs.url(filename)
            Beneficiary_DetailsSheet = pd.read_excel(myfile, sheet_name='UserDetails')
            CategoriesSheet = pd.read_excel(myfile, sheet_name='Categories')
            BrandsSheet = pd.read_excel(myfile, sheet_name='Brands')
            dbframe = Beneficiary_DetailsSheet
            for dbframe in dbframe.itertuples():
                try:
                    obj = User.objects.get(email=dbframe.Email)
                    
                  
                except User.DoesNotExist:  # or the generic "except ObjectDoesNotExist:"
                    is_active = '0'
                    my_random = random_string(8)
                    unique_user_id = '#1'+my_random
                    
                    user=User.objects.create_user(email=dbframe.Email, username=dbframe.Username,
                               first_name=dbframe.Firstname, last_name=dbframe.Lastname,
                               password='welcome',is_active=is_active
                               )
                    UserData.objects.create(user=user,unique_user_id=unique_user_id)
                    user = User.objects.get(email=dbframe.Email)
                    token = RefreshToken.for_user(user).access_token
                    token.set_exp(lifetime=timedelta(days=36500))
                    current_site = get_current_site(request).domain
                    relativeLink = reverse('email-verify')
                    absurl = 'http://' + current_site + relativeLink +'?token=' + str(token)
                    reciver_mail = user.email
                    message = MIMEMultipart()
                    message['From'] = sender_address
                    message['To'] = reciver_mail
                    message['Subject'] = 'Registration confirmation! '
                    mail_content = 'hello' + ' ' + user.username + ' please click this below  link to verify your account ' \
                        '\n ' + absurl
                    message.attach(MIMEText(mail_content, 'plain'))
                    s = smtplib.SMTP('smtp.gmail.com', 587)
                    s.starttls()
                    s.login(sender_address, sender_pass)
                    text = message.as_string()
                    s.sendmail(sender_address, reciver_mail, text)
                    #return render(request,'superadmin/import_excel.html',{'message':'File uploaded Successfully'})
                    
            category = CategoriesSheet
            for category in category.itertuples():
                try:
                    catobj = Category.objects.get(category_name=category.category_name)
                except Category.DoesNotExist:  # or the generic "except ObjectDoesNotExist:"
                    my_random = random_string(8)
                    category_code='#C1'+my_random
                    
                    catobj=Category.objects.create(category_name=category.category_name,description=category.description,
                                    category_type=category.category_type,category_code=category_code         
                               )
            brand = BrandsSheet
            for brand in brand.itertuples():
                try:
                    braobj = Brands.objects.get(brand_name=brand.brand_name)
                except Brands.DoesNotExist:  # or the generic "except ObjectDoesNotExist:"
                    my_random = random_string(8)
                    brand_code='#B1'+my_random
                    
                    braobj=Brands.objects.create(brand_name=brand.brand_name,description=brand.description,brand_code=brand_code)   
                               
                    
      return JsonResponse({'message':'File Uploaded Successfully'})
  
  
  
